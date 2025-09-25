from typing import Dict, List
from openpyxl import load_workbook

excel_mappings: Dict[str, Dict[str, str]] = {}

def get_excel_mapping() -> Dict[str, Dict[str, str]]:
    """Henter mapping fra regneark"""
    global excel_mappings
    if not excel_mappings:
        raise ValueError("excel-mapping er ikke indlæst, brug load_excel_mapping først")
    return excel_mappings


def load_excel_mapping(file_path: str):    
    global excel_mappings
    
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(file_path)
        result = {}
        for worksheet in workbook.worksheets:
            mapping = {}
            for row in worksheet.iter_rows(min_row=2, max_col=2, values_only=True):
                key, value = row[0], row[1]
                if key is not None and value is not None:
                    mapping[str(key)] = str(value)
            result[worksheet.title] = mapping

        excel_mappings = result
    except Exception as e:
        raise RuntimeError(
            f"Failed to load mapping from Excel file '{file_path}': {str(e)}"
        ) from e